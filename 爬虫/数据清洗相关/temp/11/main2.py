# import xlrd
import json
from http.cookies import SimpleCookie

import openpyxl
from pathlib import Path
# from tqdm import tqdm
import requests
import xlrd
import pandas as pd

def extract_product_ids(file_path):
    # 打开Excel文件
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active  # 获取活动表（默认第一个表）

    # 使用集合来存储商品ID，自动去重
    product_ids = set()

    # 遍历工作表中的每一行，从第二行开始（假设第一行是表头）
    for row_idx in range(2, sheet.max_row + 1):  # sheet.max_row 给出总行数
        # 假设商品 ID 在第一列
        product_id = sheet.cell(row=row_idx, column=5).value  # 获取第一列的值

        if product_id:  # 确保商品 ID 不是空值
            print(f"原始商品 ID：'{product_id}'")  # 打印原始商品ID，帮助调试
            # 拆分商品 ID 字符串（按换行符 \n）
            product_id_list = str(product_id).split("\n")
            for id in product_id_list:
                cleaned_id = id.strip()  # 去除每个商品ID的空格
                if cleaned_id:  # 如果清理后的ID不为空
                    product_ids.add(cleaned_id)

    # 如果需要返回列表格式
    return list(product_ids)

def save_to_json(data, file_path):
    # 将商品ID列表转换为逗号分隔的字符串
    data_str = ",".join(data)  # 使用逗号将商品ID连接成一个字符串
    # 打印调试信息，确保字符串没有换行符
    print("商品ID字符串:", repr(data_str))  # 使用repr打印字符串，以便查看是否有换行符

    # 将数据保存为JSON格式
    with open(file_path, 'w') as json_file:
        # 使用json.dump()时，设置ensure_ascii=False，以确保正确处理非ASCII字符
        json.dump(data_str, json_file, ensure_ascii=False)

def list_files_in_directory(directory_path):
    path = Path(directory_path)
    return [file for file in path.iterdir() if file.is_file()]

if __name__ == '__main__':
    # directory_path = '26'  # 替换为你的目录路径
    # output_path = 'josn'
    #
    # all_files = list_files_in_directory(directory_path)
    #
    # file_paths = all_files  # 文件路径列表
    #
    # for file_path in file_paths:
    #     str_id = file_path.stem  # 获取文件名，不包括扩展名
    #     json_output_path = f'{output_path}/product_ids_{str_id}.json'
    #
    #     unique_product_ids = extract_product_ids(file_path)
    #     save_to_json(unique_product_ids, json_output_path)
    #
    #     print(f"商品ID已保存到 {json_output_path}")

    unique_product_ids1 = extract_product_ids("test1.xlsx")
    save_to_json(unique_product_ids1, "product_ids_text1.json")
    unique_product_ids2 = extract_product_ids("test2.xlsx")
    save_to_json(unique_product_ids2, "product_ids_text2.json")
    all_product_ids = set(unique_product_ids1 + unique_product_ids2)  # 合并两个列表并自动去重

    # 合并两个文件中的商品 ID
    all_product_ids = set(unique_product_ids1 + unique_product_ids2)  # 合并两个列表并自动去重

    # 打印每个商品 ID
    for product_id in sorted(all_product_ids):  # 按字母顺序打印
        print(product_id)
    #
    # url = 'http://219b3c385a823369.pdd4.myjjing.com/api/admin/goods/v2/list_v2'
    # # rawdata为token
    # rawdata = 'source_data=%7B%22person%22:%22%E5%88%98%E4%B8%B0%E6%B4%81%22%7D; xd-metadata=shop_id=66e29a8543246d8d9e56ad13; sajssdk_2015_new_user_219b3c385a823369_pdd4_myjjing_com=1; sa_jssdk_2015_219b3c385a823369_pdd4_myjjing_com=%7B%22distinct_id%22%3A%22%E7%9C%9F%E7%BB%B4%E6%96%AFJEANSWEST%E7%91%9E%E6%AC%A2%E7%91%B6%E4%B8%93%E5%8D%96%E5%BA%97%3Apdd74578728430%22%2C%22first_id%22%3A%221923b7498741c0-0266f63cd26b2d-4c657b58-1328640-1923b749875197d%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E5%BC%95%E8%8D%90%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC%22%2C%22%24latest_referrer%22%3A%22https%3A%2F%2Fcustomer.xiaoduoai.com%2F%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTkyM2I3NDk4NzQxYzAtMDI2NmY2M2NkMjZiMmQtNGM2NTdiNTgtMTMyODY0MC0xOTIzYjc0OTg3NTE5N2QiLCIkaWRlbnRpdHlfbG9naW5faWQiOiLnnJ%2Fnu7Tmlq9KRUFOU1dFU1TnkZ7mrKLnkbbkuJPljZblupc6cGRkNzQ1Nzg3Mjg0MzAifQ%3D%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%24identity_login_id%22%2C%22value%22%3A%22%E7%9C%9F%E7%BB%B4%E6%96%AFJEANSWEST%E7%91%9E%E6%AC%A2%E7%91%B6%E4%B8%93%E5%8D%96%E5%BA%97%3Apdd74578728430%22%7D%7D; isGoodsDetailBackGoodsList=false; SID=2|1:0|10:1727584439|3:SID|44:MjBjMWMxYzQ2ZjNiNGExYThlMWNkZmE1ZmViNzhlMjg=|de5c4035032aaa2d2f61f1584e10a4a2993b5fc68dc8d46b0b4375c380bf12bf'
    # cookie_jar = SimpleCookie()
    # cookie_jar.load(rawdata)
    # cokies = {}
    # for key, morsel in cookie_jar.items():
    #     cokies[key] = morsel.value
    #
    # # print(goods_ids)
    # param = json.loads('''{
    #    "goods_category_id": "",
    #    "bound_to_chart": "all",
    #    "bound_to_activity": "",
    #    "keyword": "162371044481",
    #    "keyword_type": "link",
    #    "sort": "update_time_desc ",
    #    "skip": 0,
    #    "limit": 10,
    #    "cid": "",
    #    "seller_cid": "",
    #    "status": 0
    #  }''')
    # flag = False  # 标志变量
    # for product_id in unique_product_ids:
    #     param["keyword"] = str(product_id)
    #     rsp = requests.post(url, json=param, cookies=cokies)
    #     j = rsp.json()
    #     if not flag:
    #         print(j)
    #         flag = True
    #     if rsp.status_code != 200:
    #         print(j)
    #         continue

    # save_to_json(unique_product_ids, "product_ids_text.json")
    #
    # print(f"商品ID已保存到")